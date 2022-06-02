import csv
import requests
from openpyxl import load_workbook
import time
import json
from urllib.request import urlopen
import requests
import datetime

today = datetime.datetime.now ().strftime ( "%Y-%m-%dT%H:%M:%S" )
url = "http://10.50.90.202:8080"
user = "Wallop2"
password = "Wa!!0p@chan2022"
severity = "NSA6"
createby = "CFM"
reportby = "01063472"


def openfromxlsx ( ) :
    readfile = load_workbook ( "D:\\AUTO\\SCCD.xlsx" )
    cur_sheet = readfile.active
    count = 0
    for row in cur_sheet :
        if not all ( [ cell.value == None
                       for cell in row
                       ] ) :
            count += 1

    print ( "line counts : " + str ( count ) )

    for i in range ( count ) :
        i += 1
        Column1 = (cur_sheet.cell ( i , 1 ).value)
        Column2 = (cur_sheet.cell ( i , 2 ).value)
        Column3 = (cur_sheet.cell ( i , 3 ).value)

        newHeaders = { 'Content-type' : 'application/json' }
        response = requests.post ( url + '/nttservice/msom-ticket' ,
                                   json = { 'externalsystem' : 'CFM' ,
                                            'tickettype' : 'Incident' ,
                                            "trueseveritydesc" : severity ,
                                            'subject' : Column1 ,
                                            'longdescription' : Column2 ,
                                            'classification' : Column3 ,
                                            'createdby' : createby ,
                                            'reportedby' : reportby ,
                                            'reportdate' : today ,
                                            'affecteddate' : today } ,
                                   headers = newHeaders , auth = (user , password) )

        print ( Column1 + "|" + Column2 + "|" + Column3 )
        print ( response.status_code )
        print ( response.content )
        time.sleep ( 2 )


def addworklogfromxlsx ( ) :
    readfile = load_workbook ( "D:\\AUTO\\SCCDLOG.xlsx" )
    cur_sheet = readfile.active
    count = 0
    for row in cur_sheet :
        if not all ( [ cell.value == None
                       for cell in row
                       ] ) :
            count += 1

    print ( "line counts : " + str ( count ) )

    for i in range ( count ) :
        i += 1
        Column1 = (cur_sheet.cell ( i , 1 ).value)
        Column2 = (cur_sheet.cell ( i , 2 ).value)
        Column3 = (cur_sheet.cell ( i , 3 ).value)

        newHeaders = { 'Content-type' : 'application/json' , "Content-Encoding" : "gzip, deflate, br" }
        response = requests.post ( url + '/nttservice/msom-ticket-worklog' ,
                                   json = { "createby" : reportby ,
                                            "recordkey" : Column1 ,
                                            "subject" : "Update log " + Column2 ,
                                            "longdescription" : Column3

                                            } ,
                                   headers = newHeaders , auth = (user , password) )

        print ( Column1 + "|" + Column2 + "|" + Column3 )
        print ( response.status_code )
        print ( response.content )
        time.sleep ( 2 )


addworklogfromxlsx ( )
print ( 'Enter to exit:' )
x = input ( )


import json
from urllib.request import urlopen
import requests
import datetime

today = datetime.datetime.now().strftime( "%Y-%m-%dT%H:%M:%S" )
url = "http://10.50.90.202:8080"
user = "Wallop2"
password = "Wa!!0p@chan2022"
ticketid = "TT202205100066"
severity = "NSA6"
classification = "MSOM-CVIP-SQ \\ PROBLEM AREA \\ NEA"
subject = "PA-BMA05-52204_พื้นที่ซอยนิคมแหลมฉบัง 3/5 แขวงทุ่งสุขลา อำเภอศรีราชา จังหวัดชลบุรี"
description = "รบกวนตรวจสอบเพื่อหาแนวทางปรับปรุงคุณภาพสัญญาณในพื้นที่ <br>Problem : จับสัญญาณ 3G/4G และเข้า Internet ได้ Low Speed บางพื้นที่<br>PA-BMA05-52204_พื้นที่ซอยนิคมแหลมฉบัง 3/5  แขวงทุ่งสุขลา อำเภอศรีราชา จังหวัดชลบุรี<br>จังหวัด :  ชลบุรี<br>อำเภอ /เขต : ศรีราชา<br>ตำบล/แขวง : ทุ่งสุขลา<br>LATITUDE : 13.076294<br>LONGITUDE: 100.908871<br>Complaints : 8"
cause = "MS014"
subcause = "MS0037"
worklog = "testing by user via python script"
wonumber = "A4660701"


def get () :
    json_url = urlopen( url + "/nttservice/msom-ticket?ticketid=" + ticketid )

    data = json.loads( json_url.read() )

    print( data )


def open () :
    newHeaders = { 'Content-type' : 'application/json' }
    response = requests.post( url + '/nttservice/msom-ticket' ,
                              json = { 'externalsystem' : 'CFM' ,
                                       'tickettype' : 'Incident' ,
                                       "trueseveritydesc" : severity ,
                                       'subject' : subject ,
                                       'longdescription' : description ,
                                       'classification' : classification ,
                                       'createdby' : 'CFM' ,
                                       'reportedby' : '01063472' ,
                                       'reportdate' : today ,
                                       'affecteddate' : today } ,
                              headers = newHeaders , auth = (user , password) )

    print( response.status_code )
    print( response.content )


def resolve () :
    newHeaders = { 'Content-type' : 'application/json' , "Content-Encoding" : "gzip, deflate, br" }
    response = requests.patch( url + '/nttservice/msom-ticket' ,
                               json = { "ticketid" : ticketid ,
                                        "status" : "RESOLVED" ,
                                        "externalsystem" : "CFM" ,
                                        "updatedby" : "01063472" ,
                                        "failurecode" : "MSOM" ,
                                        "problemcode" : cause ,
                                        "causecode" : subcause ,
                                        "remedycode" : "MS01" ,
                                        "restorationdate" : today
                                        } ,
                               headers = newHeaders , auth = (user , password) )

    print( response.status_code )
    print( response.content )


def addworklog () :
    newHeaders = { 'Content-type' : 'application/json' , "Content-Encoding" : "gzip, deflate, br" }
    response = requests.post( url + '/nttservice/msom-ticket-worklog' ,
                              json = { "createby" : "01063472" ,
                                       "recordkey" : ticketid ,
                                       "subject" : "Update log" + ticketid ,
                                       "longdescription" : worklog

                                       } ,
                              headers = newHeaders , auth = (user , password) )

    print( response.status_code )
    print( response.content )


def assignactivity () :
    newHeaders = { 'Content-type' : 'application/json' , "Content-Encoding" : "gzip, deflate, br" }
    response = requests.post( url + '/nttservice/msom-activity' ,
                              json = { "ticketid" : ticketid ,
                                       "externalsystem" : "CFM" ,
                                       "ciname" : "CCS0001A" ,
                                       "subject" : "Test subject" ,
                                       "assignedactivityowner" : "UPC01E-RF"

                                       } ,
                              headers = newHeaders , auth = (user , password) )

    print( response.status_code )
    print( response.content )


def statusactivity () :
    newHeaders = { 'Content-type' : 'application/json' , "Content-Encoding" : "gzip, deflate, br" }
    response = requests.patch( url + '/nttservice/msom-activity' ,
                               json = { "wonum" : wonumber ,
                                        "externalsystem" : "CFM" ,
                                        "status" : "INPRG" ,
                                        "updateby" : "01063472"

                                        } ,
                               headers = newHeaders , auth = (user , password) )

    print( response.status_code )
    print( response.content )


def completeactivity () :
    newHeaders = { 'Content-type' : 'application/json' , "Content-Encoding" : "gzip, deflate, br" }
    response = requests.patch( url + '/nttservice/msom-activity' ,
                               json = { "wonum" : wonumber ,
                                        "externalsystem" : "CFM" ,
                                        "status" : "COMP" ,
                                        "updateby" : "01063472" ,
                                        "failurecode" : "MSOM" ,
                                        "problemcode" : "MS014" ,
                                        "causecode" : "MS0037" ,
                                        "remedycode" : "MS01" ,
                                        "restorationdate" : today

                                        } ,
                               headers = newHeaders , auth = (user , password) )

    print( response.status_code )
    print( response.content )

def addworklogactivity () :
    newHeaders = { 'Content-type' : 'application/json' , "Content-Encoding" : "gzip, deflate, br" }
    response = requests.patch( url + '/nttservice/msom-activity' ,
                               json = { "wonum" : wonumber ,
                                        "externalsystem" : "CFM" ,
                                        "status" : "COMP" ,
                                        "updateby" : "01063472" ,
                                        "failurecode" : "MSOM" ,
                                        "problemcode" : "MS014" ,
                                        "causecode" : "MS0037" ,
                                        "remedycode" : "MS01" ,
                                        "restorationdate" : today

                                        } ,
                               headers = newHeaders , auth = (user , password) )

    print( response.status_code )
    print( response.content )

print( ticketid + " " + today )
